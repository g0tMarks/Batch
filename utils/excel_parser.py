import openpyxl
import logging
from typing import List, Dict, Any, Optional

logger = logging.getLogger(__name__)

class ExcelParsingError(Exception):
    """Custom exception for Excel parsing errors"""
    pass

def validate_student_data(student: Dict[str, Any]) -> bool:
    """
    Validate that a student record has all required fields and they are not empty
    
    Args:
        student: Dictionary containing student data
        
    Returns:
        bool: True if student data is valid, False otherwise
    """
    required_fields = [
        'student_name',
        'year',
        'gender',
        'adjectives',
        'academic_performance',
        'extracurricular_activities',
        'other',
        'sample_report'
    ]
    
    # Check if all required fields exist and are not empty
    for field in required_fields:
        if field not in student or not student[field]:
            logger.warning(f"Missing or empty required field '{field}' for student {student.get('student_name', 'unknown')}")
            return False
            
        
    # Validate gender is one of the expected values
    valid_genders = ['male', 'female', 'other', 'm', 'f', 'o', 'M', 'F', 'O']
    if student['gender'].lower() not in valid_genders:
        logger.warning(f"Invalid gender '{student['gender']}' for student {student['student_name']}")
        return False
        
    return True

def read_student_data_from_excel(file_path: str, sheet_name: str = "Sheet1") -> List[Dict[str, Any]]:
    """
    Read and validate student data from an Excel file
    
    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet to read (default: "Sheet1")
        
    Returns:
        List[Dict[str, Any]]: List of valid student records
        
    Raises:
        ExcelParsingError: If there are issues reading or parsing the file
    """
    try:
        logger.info(f"Reading Excel file: {file_path}")
        workbook = openpyxl.load_workbook(file_path)
        
        if sheet_name not in workbook.sheetnames:
            raise ExcelParsingError(f"Sheet '{sheet_name}' not found in workbook")
            
        sheet = workbook[sheet_name]
        
        # Define expected headers in both formats
        # Note: Both 'Student Name'/'student_name' and 'Student Number'/'student_number' map to 'student_name'
        # Matching is case-insensitive, so all variations are handled
        expected_headers = {
            'Student Name': 'student_name',
            'Student Number': 'student_name',  # Accept student_number as alternative
            'Year': 'year',
            'Gender': 'gender',
            'Adjectives': 'adjectives',
            'Academic Performance': 'academic_performance',
            'Extracurricular Activities': 'extracurricular_activities',
            'Other': 'other',
            'Sample Report': 'sample_report'
        }
        
        # Get headers from the file
        headers = [str(cell.value).strip() if cell.value else '' for cell in sheet[1]]
        
        # Create a mapping of actual headers to expected format
        header_mapping = {}
        required_normalized_fields = {'student_name', 'year', 'gender', 'adjectives', 
                                     'academic_performance', 'extracurricular_activities', 
                                     'other', 'sample_report'}
        
        for header in headers:
            # Try to match the header in either format (case-insensitive)
            for expected, normalized in expected_headers.items():
                # Match against both the expected header and the normalized value
                # This handles 'Student Name', 'student_name', 'Student Number', 'student_number', etc.
                if header.lower() == expected.lower() or header.lower() == normalized.lower():
                    # Only add if we don't already have this normalized field
                    # (to handle cases where both 'Student Name' and 'Student Number' are present)
                    if normalized not in {v for v in header_mapping.values()}:
                        header_mapping[header] = normalized
                    break
        
        # Check if we have all required normalized fields
        mapped_normalized_fields = set(header_mapping.values())
        if mapped_normalized_fields != required_normalized_fields:
            missing_fields = required_normalized_fields - mapped_normalized_fields
            raise ExcelParsingError(f"Missing required headers. Missing: {missing_fields}")
        
        valid_students = []
        skipped_rows = []
        
        for row in range(2, sheet.max_row + 1):
            try:
                # Create student dictionary using the header mapping
                student = {}
                for col, header in enumerate(headers, 1):
                    if header in header_mapping:
                        student[header_mapping[header]] = sheet.cell(row=row, column=col).value
                
                # Skip empty rows
                if all(value is None for value in student.values()):
                    logger.debug(f"Skipping empty row {row}")
                    continue
                    
                # Validate student data
                if validate_student_data(student):
                    valid_students.append(student)
                    logger.debug(f"Added valid student data for {student['student_name']}")
                else:
                    skipped_rows.append(row)
                    logger.warning(f"Skipping invalid student data in row {row}")
                    
            except Exception as e:
                logger.error(f"Error processing row {row}: {str(e)}")
                skipped_rows.append(row)
                
        if skipped_rows:
            logger.warning(f"Skipped {len(skipped_rows)} rows due to invalid data: {skipped_rows}")
            
        if not valid_students:
            raise ExcelParsingError("No valid student data found in the file")
            
        logger.info(f"Successfully parsed {len(valid_students)} valid student records")
        return valid_students
        
    except Exception as e:
        logger.error(f"Error reading Excel file: {str(e)}")
        raise ExcelParsingError(f"Failed to read Excel file: {str(e)}")
